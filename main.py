import pygame
import random
import math
from random import *
from openpyxl import Workbook
from time import sleep
# Images from www.flaticon.com

# intialize the pygame
pygame.init()

# Screen/Window
screen = pygame.display.set_mode((800, 850))

# Title and Icon
pygame.display.set_caption("704 Machine Learning")
icon = pygame.image.load('car.png')
pygame.display.set_icon(icon)

# Player
playerImage = pygame.image.load('playercar.png')
playerX = 375
playerY = 750
playerXSpeed = 0
leftSpeed = -0.6
rightSpeed = 0.6

# Obstacle
obstacleImage = pygame.image.load('cone.png')
obstacleX = randint(200, 600)
obstacleY = 0
obstacleYSpeed = 0.8

# Coins
coinImage = pygame.image.load('coin.png')
coinX = randint(200, 600)
coinY = 0
coinSpeed = 1

# Score
scoreValue = 0
font = pygame.font.Font('freesansbold.ttf', 26)
textX = 10
textY = 10

# Timer
TIME = pygame.USEREVENT + 1
pygame.time.set_timer(TIME, 1000)
time = 0
# Game Over
gameOverFont = pygame.font.Font('freesansbold.ttf', 64)

# Flavour
roadMarking1X = 400
roadMarking1Y = 0
roadMarking2X = 400
roadMarking2Y = 425
roadMarkingSpeed = 0.8

# Error rate 24 and it collides
errorSize = 50
errorRate = 1

# Game restart
gameOver = False

# Obstacle near
nearBool = 0

### Data sending ###

wb = Workbook()
# Grab the active worksheet
ws = wb.active

# Set up top titles in excel
ws['A1'] = "PlayerPos"
ws['B1'] = "ObstacleXPos"
ws['C1'] = "ObstacleY"
ws['D1'] = "CoinXPos"
ws['E1'] = "CoinYPos"
ws['F1'] = "ObstacleNear"

wb.save('MLDrivingData.csv')

###############################################################################################################
## Functions ##

# Renders text(score) to screen at x , y
def showScore(x, y):
    score = font.render("Score: " + str(scoreValue), True, (0, 0, 0))
    screen.blit(score, (x, y))


# Renders game over text(score) to screen at x , y
def gameOverText():
    gameOverText = gameOverFont.render("Score: " + str(scoreValue), True, (0, 0, 0))
    screen.blit(gameOverText, (280, 300))


# Draw player at x , y
def player(x, y):
    screen.blit(playerImage, (x, y))


# Draw obstacle at x , y
def obstacle(x, y):
    screen.blit(obstacleImage, (x, y))


# Draw coin at x , y
def coin(x, y):
    screen.blit(coinImage, (x, y))


# distance between 2 coordinates, (less than 27), collision
def isCollision(obstacleX, obstacleY, playerX, playerY):
    distance = math.sqrt((math.pow(obstacleX - playerX, 2)) + (math.pow(obstacleY - playerY, 2)))
    if distance < 28:
        return True
    else:
        return False

##############################################################################################################

# Game Loop #

running = True
while running:

    # Draw background, grass , barriers and road markings
    screen.fill((75, 75, 75))
    pygame.draw.rect(screen, (100, 140, 100), [0, 0, 150, 850])
    pygame.draw.rect(screen, (30, 30, 30), [140, 0, 10, 850])
    pygame.draw.rect(screen, (100, 140, 100), [650, 0, 200, 850])
    pygame.draw.rect(screen, (30, 30, 30), [650, 0, 10, 850])
    pygame.draw.rect(screen, (200, 200, 200), [roadMarking1X, roadMarking1Y, 20, 100])
    pygame.draw.rect(screen, (200, 200, 200), [roadMarking2X, roadMarking2Y, 20, 100])

    # bool for when the obstacle is near bottom (data sending)
    if obstacleY >= 600:
        nearBool = 1
    elif obstacleY < 600:
        nearBool = 0



    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False

        # Player input
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_LEFT:
                playerXSpeed = leftSpeed

            if event.key == pygame.K_RIGHT:
                playerXSpeed = rightSpeed

            # Player Speed up slow down logic
            if event.key == pygame.K_UP:
                obstacleYSpeed += 0.1
                roadMarkingSpeed += 0.1
                coinSpeed += 0.1
            if event.key == pygame.K_DOWN:
                obstacleYSpeed -= 0.1
                roadMarkingSpeed -= 0.1
                coinSpeed -= 0.1

        elif event.type == TIME:
            time += 1

        if event.type == pygame.KEYUP:
            if event.key == pygame.K_LEFT or event.key == pygame.K_RIGHT:
                playerXSpeed = 0

    # some form of ai thing

    if (obstacleY > 400) and (obstacleX <= playerX <= obstacleX + errorSize):
        playerXSpeed = rightSpeed
    elif coinX <= playerX:
        playerXSpeed = leftSpeed
    if (obstacleY > 400) and obstacleX >= playerX >= obstacleX - errorSize:
        playerXSpeed = leftSpeed
    elif coinX >= playerX:
        playerXSpeed = rightSpeed

    # Extra data for obstacle the nearer the bottom
    if (obstacleY >= 450) and (obstacleY <= 450.7):
        ws.append([playerX, obstacleX, obstacleY, coinX, coinY,nearBool])
    if (obstacleY >= 550) and (obstacleY <= 550.7):
        ws.append([playerX, obstacleX, obstacleY, coinX, coinY,nearBool])
    if (obstacleY >= 650) and (obstacleY <= 650.7):
        ws.append([playerX, obstacleX, obstacleY, coinX, coinY,nearBool])
    if (obstacleY >= 750) and (obstacleY <= 750.7):
        ws.append([playerX, obstacleX, obstacleY, coinX, coinY,nearBool])

    # Update player left/right speed
    playerX += playerXSpeed

    # Checking for boundaries(barriers)
    if playerX <= 100:
        playerX = 100
    elif playerX > 640:
        playerX = 640

    # Obstacle movement + re spawn
    obstacleY += obstacleYSpeed

    if obstacleY >= 850:
        ws.append([playerX, obstacleX, obstacleY, coinX, coinY, nearBool])
        obstacleY = -50
        obstacleX = randint(175, 625)
        scoreValue += 1

        # Decrease error size by chance so it fails at some point and can restart
        x = randint(1, 100)
        if x >=75:
            errorSize = errorSize - errorRate

    # Road markings + re spawn #
    roadMarking1Y += roadMarkingSpeed
    roadMarking2Y += roadMarkingSpeed
    if roadMarking1Y >= 850:
        roadMarking1Y = -50
        roadMarking1X = 400
    if roadMarking2Y >= 850:
        roadMarking2Y = -50
        roadMarking2X = 400

    # Coin logic #
    coinY += coinSpeed
    coinCollision = isCollision(coinX, coinY, playerX, playerY)

    if coinCollision:
        ws.append([playerX, obstacleX,obstacleY, coinX, coinY,nearBool])

        coinY = -50
        coinX = randint(175, 625)
        scoreValue += 1

    if coinY >= 850:
        ws.append([playerX, obstacleX, obstacleY, coinX, coinY, nearBool])
        coinY = -50
        coinX = randint(175, 625)

    # Collision/ Game over #
    collision = isCollision(obstacleX, obstacleY, playerX, playerY)
    if collision:
        obstacleYSpeed = 0
        playerXSpeed = 0
        roadMarkingSpeed = 0
        coinSpeed = 0
        gameOverText()
        gameOver = True

    if gameOver == True:
        obstacleY = -50
        obstacleX = randint(175, 625)
        obstacleYSpeed = 0.8
        roadMarkingSpeed = 0.8
        coinSpeed = 1
        coinY = -50
        coinX = randint(175, 625)
        scoreValue = 0
        errorSize = 50
        sleep(1)
        # Save the excel sheet(do at end to save pc hdd from dying)
        wb.save('MLDrivingData.csv')
        time = 0
        gameOver = False

    # Update image locations
    player(playerX, playerY)
    obstacle(obstacleX, obstacleY)
    coin(coinX, coinY)
    showScore(textX, textY)
    pygame.display.update()
